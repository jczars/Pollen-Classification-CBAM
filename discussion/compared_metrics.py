import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import argparse

def compare_metrics(input_file):
    # Carregar a planilha
    df = pd.read_excel(input_file, sheet_name="comparar")
    
    # Criar DataFrames para cada base
    orig_400 = df[df['Base'] == 'Orig']
    eq_400 = df[df['Base'] == 'EQ']
    pl_400 = df[df['Base'] == 'PL']
    
    # Função para comparar métricas
    def compare(df1, df2):
        return df1.merge(df2, on='Classes', suffixes=('_orig', '_comp'))
    
    # Comparações
    comparison_eq = compare(orig_400, eq_400)
    comparison_pl = compare(orig_400, pl_400)
    
    # Criar um escritor de planilhas
    with pd.ExcelWriter(input_file, engine='openpyxl', mode="a") as writer:
        comparison_eq.to_excel(writer, sheet_name='Comparison_Orig_EQ', index=False)
        comparison_pl.to_excel(writer, sheet_name='Comparison_Orig_PL', index=False)
    
    # Destacar valores melhorados/neutros em negrito
    wb = load_workbook(input_file)
    for sheet_name in ['Comparison_Orig_EQ', 'Comparison_Orig_PL']:
        ws = wb[sheet_name]
        header = {cell.value: cell.column for cell in ws[1]}
        
        for row in range(2, ws.max_row + 1):
            for metric in ['Accuracy', 'Precision', 'Recall', 'F1-Score']:
                col_comp = header[f'{metric}_comp']
                col_orig = header[f'{metric}_orig']
                
                if ws.cell(row, col_comp).value >= ws.cell(row, col_orig).value:  # Se melhorou ou manteve
                    ws.cell(row, col_comp).font = Font(bold=True)
    
    wb.save(input_file)
    print(f'✅ Comparação salva e destacada em {input_file}')

# Exemplo de uso
#input_file = "discussion/Comparar_literatura_B23.xlsx"  # Nome do arquivo de entrada
#compare_metrics(input_file)

# Main function to handle command-line arguments
if __name__ == "__main__":
    # Default file path configuration
    default_path = 'discussion/Comparar_literatura_CPD1_A200.xlsx'
    
    # Set up argparse to handle command-line arguments
    parser = argparse.ArgumentParser(description="Run the pollen classification process.")
    parser.add_argument(
        '--path', 
        type=str, 
        default=default_path, 
        help="Path to the workbook. If not provided, the default path will be used."
    )

    args = parser.parse_args()
    
    # Check if the given path exists
    if not os.path.exists(args.path):
        print(f"Warning: The provided workbook path '{args.path}' does not exist.")
        print("Using default workbook path.")
        args.path = default_path

    # Perform comparative analysis and Wilcoxon test
    compare_metrics(args.path)


    # Display completion message
    print(f"The results and their interpretations have been saved in the new 'results' sheet of the spreadsheet: {args.path}")