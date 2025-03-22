import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Função para carregar e filtrar um CSV
def load_and_filter_csv(file_path):
    # Carregar o CSV
    df = pd.read_csv(file_path, sep=";", decimal=",")
    
    # Filtrar classes com Support > 0
    df_filtered = df[df['Support'] > 0]
    
    # Converter colunas numéricas para float (se necessário)
    numeric_columns = ['Accuracy', 'Precision', 'Recall', 'F1-Score', 'Support']
    df_filtered[numeric_columns] = df_filtered[numeric_columns].astype(float)
    
    return df_filtered

# Função para comparar duas tabelas e formatar
def compare_and_format(df_orig, df_comp, sheet_name, output_path):
    # Garantir que ambas as tabelas tenham as mesmas classes
    common_classes = df_orig["Classes"].isin(df_comp["Classes"])
    df_orig_filtered = df_orig[common_classes]
    df_comp_filtered = df_comp[df_comp["Classes"].isin(df_orig["Classes"])]

    if df_orig_filtered.empty or df_comp_filtered.empty:
        print(f"Nenhuma classe em comum encontrada para {sheet_name}.")
        return

    # Criar um DataFrame vazio para armazenar os resultados
    comparison_df = pd.DataFrame()
    comparison_df["Classes"] = df_orig_filtered["Classes"]

    # Definir as métricas renomeadas
    metric_mapping = {
        "Accuracy": "Acc",
        "Precision": "P",
        "Recall": "R",
        "F1-Score": "F1"
    }

    # Adicionar métricas da base original
    for metric, short_name in metric_mapping.items():
        orig_metric = f"{short_name}_Orig"
        comparison_df[orig_metric] = df_orig_filtered[metric].round(3)

    # Adicionar métricas da base consolidada
    for metric, short_name in metric_mapping.items():
        comp_metric = f"{short_name}_Comp"
        comparison_df[comp_metric] = df_comp_filtered[metric].round(3)

    # Reorganizar as colunas: Classes, métricas Originais, métricas Consolidadas
    columns_order = ["Classes"]
    for metric in metric_mapping.values():
        columns_order.append(f"{metric}_Orig")
    for metric in metric_mapping.values():
        columns_order.append(f"{metric}_Comp")

    comparison_df = comparison_df[columns_order]

    # Salvar o DataFrame no Excel
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        comparison_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar formatação no Excel
    workbook = load_workbook(output_path)
    worksheet = workbook[sheet_name]

    # Definição de estilos
    equal_style = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cinza
    better_style = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro
    worse_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Vermelho

    for row in range(2, len(comparison_df) + 2):  # Começa na linha 2 para ignorar cabeçalhos
        for col_idx, metric in enumerate(metric_mapping.values(), start=1):
            orig_col = col_idx + 1  # Coluna do valor original (depois de "Classes")
            view_col = orig_col + len(metric_mapping)  # Coluna do valor da vista
            
            orig_value = worksheet.cell(row=row, column=orig_col).value
            view_value = worksheet.cell(row=row, column=view_col).value

            if isinstance(orig_value, (int, float)) and isinstance(view_value, (int, float)):
                if view_value > orig_value:
                    worksheet.cell(row=row, column=view_col).fill = better_style
                elif view_value < orig_value:
                    worksheet.cell(row=row, column=view_col).fill = worse_style
                else:
                    worksheet.cell(row=row, column=view_col).fill = equal_style

    workbook.save(output_path)
    workbook.close()
    print(f"Planilha {sheet_name} formatada com sucesso: {output_path}")

# Função principal
def main():
    # Caminhos dos arquivos CSV
    original_path = "/media/jczars/4C22F02A22F01B22/Pollen-Classification-CBAM/results/phase1/AT_densenet+cbam_exp/5_DenseNet201_reports_consolidated/class_report_test_5_DenseNet201.csv"
    equatorial_path = "/media/jczars/4C22F02A22F01B22/Pollen-Classification-CBAM/results/phase3/CPD1_TEST_A200/5_DenseNet201_EQUATORIAL_consolidated/class_report_test_5_DenseNet201.csv"
    polar_path = "/media/jczars/4C22F02A22F01B22/Pollen-Classification-CBAM/results/phase3/CPD1_TEST_A200/5_DenseNet201_POLAR_consolidated/class_report_test_5_DenseNet201.csv"

    # Caminho do arquivo Excel de saída
    excel_path = "discussion/Comparisons_build.xlsx"

    # Carregar e filtrar os DataFrames
    original_df = load_and_filter_csv(original_path)
    equatorial_df = load_and_filter_csv(equatorial_path)
    polar_df = load_and_filter_csv(polar_path)

    # Salvar os DataFrames originais no Excel
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        original_df.to_excel(writer, sheet_name="Original", index=False)
        equatorial_df.to_excel(writer, sheet_name="Equatorial", index=False)
        polar_df.to_excel(writer, sheet_name="Polar", index=False)

    print(f"DataFrames salvos no arquivo Excel: {excel_path}")

    # Comparar Original x Equatorial
    compare_and_format(original_df, equatorial_df, "Original vs Equatorial", excel_path)
    print("Comparação Original x Equatorial:")

    # Comparar Original x Polar
    compare_and_format(original_df, polar_df, "Original vs Polar", excel_path)
    print("Comparação Original x Polar:")

    print("Processo concluido com sucesso.")
if __name__ == "__main__":
    main()